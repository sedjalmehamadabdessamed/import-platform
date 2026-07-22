# -*- coding: utf-8 -*-
"""
تطبيق إدارة بيانات التصدير الجمركي - المرحلة الأولى
=====================================================
المتطلبات:
    pip install streamlit openpyxl

طريقة التشغيل:
    streamlit run app.py

ملاحظة مهمة:
    يجب أن يكون ملف "products.xlsx" في نفس مجلد هذا الملف،
    ويحتوي على عمودين بالضبط بهذا الترتيب في الصف الأول (رأس الجدول):
        اسم المنتج  |  رمز الجمارك
"""

import os
import io
import sqlite3

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# =========================================================
# إعدادات عامة للصفحة
# =========================================================
st.set_page_config(
    page_title="نظام إدارة بيانات التصدير",
    page_icon="📦",
    layout="wide",
)

# تفعيل الاتجاه من اليمين لليسار (واجهة عربية)
st.markdown(
    """
    <style>
        html, body, [class*="css"] {
            direction: rtl;
            text-align: right;
        }
        .stTextInput > div > div > input,
        .stNumberInput input,
        .stSelectbox div[data-baseweb="select"] {
            text-align: right;
            direction: rtl;
        }
        section[data-testid="stSidebar"] { direction: rtl; text-align: right; }
        div[data-testid="stMetricValue"] { direction: ltr; }
    </style>
    """,
    unsafe_allow_html=True,
)

DB_PATH = "app_database.db"
PRODUCTS_FILE = "products.xlsx"
LOGO_DIR = "app_data"


# =========================================================
# قاعدة البيانات (SQLite)
# =========================================================
@st.cache_resource
def get_connection():
    """إنشاء اتصال واحد بقاعدة البيانات يُعاد استخدامه طوال عمل التطبيق."""
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    return conn


def init_db():
    """إنشاء الجداول إذا لم تكن موجودة."""
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS company_info (
            id INTEGER PRIMARY KEY,
            name TEXT,
            commercial_register TEXT,
            tax_number TEXT,
            logo_path TEXT
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS cart_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_name TEXT,
            customs_code TEXT,
            origin_country TEXT,
            quantity REAL,
            unit TEXT,
            price REAL,
            currency TEXT
        )
        """
    )
    conn.commit()


def load_company_info():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT name, commercial_register, tax_number, logo_path FROM company_info WHERE id = 1"
    )
    row = cur.fetchone()
    if row:
        return {
            "name": row[0] or "",
            "commercial_register": row[1] or "",
            "tax_number": row[2] or "",
            "logo_path": row[3] or "",
        }
    return {"name": "", "commercial_register": "", "tax_number": "", "logo_path": ""}


def save_company_info(name, commercial_register, tax_number, logo_path):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO company_info (id, name, commercial_register, tax_number, logo_path)
        VALUES (1, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            name = excluded.name,
            commercial_register = excluded.commercial_register,
            tax_number = excluded.tax_number,
            logo_path = excluded.logo_path
        """,
        (name, commercial_register, tax_number, logo_path),
    )
    conn.commit()


def add_product_to_cart(product_name, customs_code, origin_country, quantity, unit, price, currency):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO cart_products
            (product_name, customs_code, origin_country, quantity, unit, price, currency)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (product_name, customs_code, origin_country, quantity, unit, price, currency),
    )
    conn.commit()


def get_cart_products():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, product_name, customs_code, origin_country, quantity, unit, price, currency
        FROM cart_products ORDER BY id
        """
    )
    rows = cur.fetchall()
    columns = ["id", "product_name", "customs_code", "origin_country", "quantity", "unit", "price", "currency"]
    return [dict(zip(columns, row)) for row in rows]


def delete_product_from_cart(product_id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM cart_products WHERE id = ?", (product_id,))
    conn.commit()


# =========================================================
# قراءة المنتجات من ملف Excel (باستخدام openpyxl)
# =========================================================
def _read_products_file(path):
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        name_idx, code_idx = None, None
        for i, h in enumerate(headers):
            if h == "اسم المنتج":
                name_idx = i
            elif h == "رمز الجمارك":
                code_idx = i
        if name_idx is None or code_idx is None:
            name_idx, code_idx = 0, 1  # احتياطي: أول عمودين

        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) <= max(name_idx, code_idx):
                continue
            name = row[name_idx]
            code = row[code_idx]
            if name is None or str(name).strip() == "":
                continue
            products.append(
                {"name": str(name).strip(), "code": str(code).strip() if code is not None else ""}
            )
        return products
    except Exception as e:
        st.error(f"⚠️ حدث خطأ أثناء قراءة ملف المنتجات: {e}")
        return []


@st.cache_data(show_spinner=False)
def _cached_products(path, mtime):
    # مُعامل mtime يجعل الذاكرة المؤقتة تتحدث تلقائيًا عند تعديل الملف
    return _read_products_file(path)


def load_products():
    if not os.path.exists(PRODUCTS_FILE):
        return []
    mtime = os.path.getmtime(PRODUCTS_FILE)
    return _cached_products(PRODUCTS_FILE, mtime)


# =========================================================
# توليد ملف Excel النهائي
# =========================================================
def generate_excel_file(company, products):
    wb = Workbook()

    # ---------- ورقة بيانات الشركة ----------
    ws1 = wb.active
    ws1.title = "بيانات الشركة"
    ws1.sheet_view.rightToLeft = True

    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    label_font = Font(bold=True, size=11)

    ws1.merge_cells("A1:C1")
    ws1["A1"] = "بيانات الشركة"
    ws1["A1"].font = title_font
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 26

    info_rows = [
        ("اسم الشركة", company.get("name", "")),
        ("السجل التجاري", company.get("commercial_register", "")),
        ("الرقم الجبائي", company.get("tax_number", "")),
    ]
    r = 3
    for label, value in info_rows:
        ws1.cell(row=r, column=1, value=label).font = label_font
        ws1.cell(row=r, column=2, value=value)
        r += 1

    logo_path = company.get("logo_path", "")
    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width = 120
            img.height = 120
            ws1.add_image(img, "D3")
        except Exception:
            pass  # في حال كانت صيغة الصورة غير مدعومة، نتجاهل إدراجها دون كسر التصدير

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 32
    ws1.column_dimensions["C"].width = 18

    # ---------- ورقة المنتجات ----------
    ws2 = wb.create_sheet("المنتجات")
    ws2.sheet_view.rightToLeft = True

    headers = ["اسم المنتج", "رمز الجمارك", "بلد المنشأ", "الكمية", "الوحدة", "السعر", "العملة"]
    header_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for r_idx, p in enumerate(products, start=2):
        values = [
            p["product_name"], p["customs_code"], p["origin_country"],
            p["quantity"], p["unit"], p["price"], p["currency"],
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    widths = [28, 16, 16, 10, 10, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =========================================================
# تشغيل التطبيق
# =========================================================
init_db()
company = load_company_info()

st.title("📦 نظام إدارة بيانات التصدير")
st.caption("المرحلة الأولى: بيانات الشركة + إضافة المنتجات + تصدير Excel")

st.divider()

# ---------------------------------------------------------
# 1) قسم بيانات الشركة
# ---------------------------------------------------------
st.header("🏢 بيانات الشركة")

col1, col2, col3 = st.columns(3)
with col1:
    company_name = st.text_input("اسم الشركة", value=company["name"])
with col2:
    commercial_register = st.text_input("السجل التجاري", value=company["commercial_register"])
with col3:
    tax_number = st.text_input("الرقم الجبائي", value=company["tax_number"])

logo_col1, logo_col2 = st.columns([2, 1])
with logo_col1:
    uploaded_logo = st.file_uploader("رفع شعار الشركة", type=["png", "jpg", "jpeg"])

logo_path = company["logo_path"]
if uploaded_logo is not None:
    os.makedirs(LOGO_DIR, exist_ok=True)
    ext = uploaded_logo.name.split(".")[-1].lower()
    logo_path = os.path.join(LOGO_DIR, f"logo.{ext}")
    with open(logo_path, "wb") as f:
        f.write(uploaded_logo.getbuffer())

with logo_col2:
    if logo_path and os.path.exists(logo_path):
        st.image(logo_path, caption="الشعار الحالي", width=120)

# حفظ بيانات الشركة تلقائيًا في قاعدة البيانات
save_company_info(company_name, commercial_register, tax_number, logo_path)

st.divider()

# ---------------------------------------------------------
# 2) قسم المنتجات
# ---------------------------------------------------------
st.header("🛒 إضافة منتج")

all_products = load_products()

if not all_products:
    st.error(
        f"⚠️ لم يتم العثور على ملف '{PRODUCTS_FILE}' أو أنه فارغ. "
        "تأكد من وجود الملف بجانب app.py وأنه يحتوي على عمودين: "
        "«اسم المنتج» و«رمز الجمارك»."
    )
else:
    search_col, _ = st.columns([2, 1])
    with search_col:
        search_query = st.text_input("🔍 ابحث عن منتج")

    if search_query.strip():
        filtered_products = [
            p for p in all_products if search_query.strip() in p["name"]
        ]
    else:
        filtered_products = all_products

    selected_name = None
    selected_code = ""

    if filtered_products:
        product_names = [p["name"] for p in filtered_products]
        selected_name = st.selectbox("اختر المنتج", product_names)
        selected_code = next(
            (p["code"] for p in filtered_products if p["name"] == selected_name), ""
        )
        st.text_input("رمز الجمارك (تلقائي)", value=selected_code, disabled=True)
    else:
        st.warning("لا توجد نتائج مطابقة لبحثك.")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        origin_country = st.text_input("بلد المنشأ")
    with c2:
        quantity = st.number_input("الكمية", min_value=0.0, value=1.0, step=1.0)
    with c3:
        unit_options = ["قطعة", "كلغ", "غرام", "لتر", "طن", "متر", "أخرى"]
        unit_choice = st.selectbox("الوحدة", unit_options)
        if unit_choice == "أخرى":
            unit = st.text_input("حدد الوحدة", key="custom_unit")
        else:
            unit = unit_choice
    with c4:
        price = st.number_input("السعر", min_value=0.0, value=0.0, step=0.01)

    currency_options = ["MAD", "USD", "EUR", "أخرى"]
    currency_choice = st.selectbox("العملة", currency_options)
    if currency_choice == "أخرى":
        currency = st.text_input("حدد العملة", key="custom_currency")
    else:
        currency = currency_choice

    if st.button("➕ إضافة المنتج", type="primary"):
        if not selected_name:
            st.warning("الرجاء اختيار منتج أولاً.")
        elif not origin_country.strip():
            st.warning("الرجاء إدخال بلد المنشأ.")
        else:
            add_product_to_cart(
                selected_name, selected_code, origin_country.strip(),
                quantity, unit, price, currency,
            )
            st.success(f"تمت إضافة المنتج «{selected_name}» بنجاح.")
            st.rerun()

st.divider()

# ---------------------------------------------------------
# جدول المنتجات المضافة
# ---------------------------------------------------------
st.header("📋 المنتجات المضافة")

cart = get_cart_products()

if not cart:
    st.info("لم تتم إضافة أي منتج بعد.")
else:
    header_cols = st.columns([3, 2, 2, 1.2, 1.2, 1.2, 1.2, 1])
    headers_labels = ["اسم المنتج", "رمز الجمارك", "بلد المنشأ", "الكمية", "الوحدة", "السعر", "العملة", ""]
    for col, label in zip(header_cols, headers_labels):
        col.markdown(f"**{label}**")

    for item in cart:
        row_cols = st.columns([3, 2, 2, 1.2, 1.2, 1.2, 1.2, 1])
        row_cols[0].write(item["product_name"])
        row_cols[1].write(item["customs_code"])
        row_cols[2].write(item["origin_country"])
        row_cols[3].write(item["quantity"])
        row_cols[4].write(item["unit"])
        row_cols[5].write(item["price"])
        row_cols[6].write(item["currency"])
        if row_cols[7].button("🗑️", key=f"delete_{item['id']}"):
            delete_product_from_cart(item["id"])
            st.rerun()

st.divider()

# ---------------------------------------------------------
# 3) توليد ملف Excel
# ---------------------------------------------------------
st.header("📤 تصدير البيانات")

if st.button("📊 إنشاء ملف Excel", type="primary"):
    if not company_name.strip():
        st.warning("الرجاء إدخال اسم الشركة قبل التصدير.")
    elif not cart:
        st.warning("لا توجد منتجات لتصديرها. الرجاء إضافة منتج واحد على الأقل.")
    else:
        company_data = {
            "name": company_name,
            "commercial_register": commercial_register,
            "tax_number": tax_number,
            "logo_path": logo_path,
        }
        excel_buffer = generate_excel_file(company_data, cart)
        st.success("تم إنشاء ملف Excel بنجاح ✅")
        st.download_button(
            label="⬇️ تحميل ملف Excel",
            data=excel_buffer,
            file_name=f"{company_name}_تصدير.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )